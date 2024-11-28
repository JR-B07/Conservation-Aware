from flask import Flask, render_template, request, send_file, url_for, redirect, jsonify, session
import pandas as pd
from pathlib import Path
import secrets

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)

data_dir = Path('data')
data_dir.mkdir(exist_ok=True)


def initialize_qfd_data(Ques, Comos):
    """Inicializa la estructura de datos completa para QFD"""
    qfd_data = {
        'Ques': Ques,
        'Comos': Comos,
        'matriz': [[0 for _ in Comos] for _ in Ques],
        'importancia': [1 for _ in Ques],
        'resultados_internos': [0 for _ in Comos],
        'importancia_total': [0 for _ in Comos],
        'importancia_relativa': [0.0 for _ in Comos]
    }
    return qfd_data


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/LosQFD')
def LosQFD():
    return render_template('LosQFD.html')


@app.route('/casa')
def casa():
    return render_template('casa.html')


@app.route('/tabla_editable')
def tabla_editable():
    qfd_data = session.get('qfd_data', {
        'Ques': [],
        'Comos': [],
        'matriz': [],
        'importancia': [],
        'resultados_internos': [],
        'importancia_total': [],
        'importancia_relativa': []
    })
    return render_template('tabla_editable.html', qfd_data=qfd_data)


@app.route('/generar_tabla', methods=['POST'])
def generar_tabla():
    try:
        Ques = [q.strip() for q in request.form.get(
            'ques', '').split(',') if q.strip()]
        Comos = [c.strip() for c in request.form.get(
            'comos', '').split(',') if c.strip()]

        if not Ques or not Comos:
            return jsonify({'error': 'Se requieren QUÉs y CÓMOs válidos'}), 400

        qfd_data = initialize_qfd_data(Ques, Comos)
        session['qfd_data'] = qfd_data

        return jsonify({
            'status': 'success',
            'redirect': url_for('tabla_editable')
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/actualizar_matriz', methods=['POST'])
def actualizar_matriz():
    try:
        datos = request.get_json()
        if not datos:
            return jsonify({'status': 'error', 'message': 'No se recibieron datos'}), 400

        # Guardar en la sesión
        session['qfd_data'] = datos
        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"Error al actualizar matriz: {str(e)}")  # Para debugging
        return jsonify({'status': 'error', 'message': str(e)}), 400


@app.route('/exportar_qfd')
def exportar_qfd():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        qfd_data = session.get('qfd_data')
        if not qfd_data:
            return "No hay datos para exportar", 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz QFD"

        # Estilos
        header_fill = PatternFill(
            start_color="198754", end_color="198754", fill_type="solid")
        light_fill = PatternFill(start_color="f8f9fa",
                                 end_color="f8f9fa", fill_type="solid")
        bold_font = Font(bold=True, color="FFFFFF")
        normal_font = Font(bold=False)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Encabezados
        ws['A1'] = "PRIORIDAD"
        ws['A1'].fill = header_fill
        ws['A1'].font = bold_font
        ws['B1'] = "Importancia"
        ws['B1'].fill = light_fill
        ws['B1'].font = bold_font

        # CÓMOs
        for col, como in enumerate(qfd_data['Comos'], start=3):
            cell = ws.cell(row=1, column=col, value=como)
            cell.fill = light_fill
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = border

        # QUÉs y valores
        for row, que in enumerate(qfd_data['Ques'], start=2):
            # QUÉ
            ws.cell(row=row, column=1, value=que).fill = light_fill
            # Importancia
            ws.cell(row=row, column=2,
                    value=qfd_data['importancia'][row-2]).alignment = center_align

            # Valores de la matriz
            for col, valor in enumerate(qfd_data['matriz'][row-2], start=3):
                cell = ws.cell(row=row, column=col, value=valor)
                cell.alignment = center_align
                cell.border = border

        # Filas de resultados
        row_resultados = len(qfd_data['Ques']) + 2
        ws.cell(row=row_resultados, column=1,
                value="Resultados internos").fill = light_fill
        ws.cell(row=row_resultados + 1, column=1,
                value="IMPORTANCIA").fill = light_fill
        ws.cell(row=row_resultados + 2, column=1,
                value="IMPORTANCIA RELATIVA").fill = light_fill

        # Valores de resultados
        for col, (res, imp, rel) in enumerate(zip(
            qfd_data['resultados_internos'],
            qfd_data['importancia_total'],
            qfd_data['importancia_relativa']
        ), start=3):
            ws.cell(row=row_resultados, column=col,
                    value=res).alignment = center_align
            ws.cell(row=row_resultados + 1, column=col,
                    value=imp).alignment = center_align
            ws.cell(row=row_resultados + 2, column=col,
                    value=f"{rel:.2f}%").alignment = center_align

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(qfd_data['Comos']) + 3):
            ws.column_dimensions[chr(64 + col)].width = 15

        # Guardar archivo
        excel_path = data_dir / 'matriz_qfd.xlsx'
        wb.save(excel_path)

        return send_file(
            excel_path,
            as_attachment=True,
            download_name='matriz_qfd.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return str(e), 500


if __name__ == '__main__':
    app.run(debug=True)
